#! python3
#Document Circulation - Excel Testing

import openpyxl
import re
import os
import subprocess

def stmts():
	"""Retrieve Statements"""
	block = ""
	found = False
	
	text_file = open("L:\\Merge and Print\\Print_Files\\" + current[0] + "\\" + current[1] + " STATEMENTS.txt")
	
	for line in text_file:
		if found:
			if WORD in line:
				if search_string in line:
					block += line
				else:
					found = False
					break
			else:
				block += line
		else:
			if search_string in line:
				found = True
				global output
				if output:
					block = line
				else:
					block = line.strip() + "\n"
					output = True
					
	output_file.write(block)
	#close the text file
	text_file.close

def process_invoices():
    """Retrieve Invoices"""
    block = ""
    found = False

    #open invoice text file
    text_file = open("L:\\Merge and Print\\Print_Files\\" + region + "\\" + agency + \
                     " INVOICES.txt")

    #parse through each line of the invoice file
    for line in text_file:
        if found:
            if WORD in line:
                if search_string in line:
                    block += line
                else:
                    found = False
                    break
            else:
                block += line
        else:
            if search_string in line:
                found = True
                global output
                if output:
                    block = line
                else:
                    block = line.strip() + "\n"
                    output = True

    #write contents to output file
    output_file.write(block)
    #print(block)
    #close text file
    text_file.close()
	
def process_credits():
	"""Retrieve credit memos"""
	block = ""
	found = False
	verification_num = ""
		
	#open credit memo text file
	text_file = open("L:\\Merge and Print\\Print_Files\\" + current[0] + "\\" + current[1] + " CREDITS.txt")
		
	#parse through each line
	for line in text_file:
		if found:
			if WORD in line:
				output_file.write(block)
				block = ""
				global output
				output = True
				if search_string in line:
					block += line
				else:
					found = False
					break
			elif "for verification:" in line.lower():
				verification_num = invoice_number
				block = ""
				found = False
				output = False
			else:
				block += line
		else:
			if search_string in line:
				found = True
				if output:
					block = line
				else:
					block = line.strip() + "\n"
					
	text_file.close()

		

WORD = "PRIMARY SEQ#:"
		
#open dealer document list
wb = openpyxl.load_workbook("L:\Merge and Print\Customer_Documents - TEST.xlsm")
sheet = wb.active

for i in range(2,sheet.max_row):
	output = False
	current = []
	for a in range(1,9):
		current = current + [sheet.cell(row=i, column=a).value]

	
	if current[6] =="y":
		account_number_str = str(current[3])
		while len(account_number_str)<12:
			account_number_str = " " + account_number_str
		
		search_string = WORD + account_number_str
		print(search_string)
		
		#creat output file
		path = "L:\Merge and Print\\In_Process\\" + current[0] + "\\" + current[1] + \
				" " + str(current[3]) + " " + re.sub('[^A-Za-z0-9]+',' ',str(current[2])) + ".txt"
		output_file = open(path,"a")

		process_credits()
		output_file.close()
		if output == False:
			os.remove(path)
		else:
			subprocess.run(["C:\\Program Files (x86)\\Text2PDF v1.5\\txt2pdf.exe",path,"-pfs8","-plm25","-prm25", \
				"-ptm25","-pbm25","-pdn:Lucida Console"])
			os.remove(path)
	
	