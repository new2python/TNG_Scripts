#! python3
#Document Circulation
#Created by Curtis Cholon
#04/24/2017
"""Processes computac documents pulled from the Merge and Print Processes.
Documents are retrieved, formatted and sent via email or fax"""

#-----------------------------------------------------------------------------------------------#
#Debugging Disabled for now - need to learn how useto... and how to write to text file.
# import logging
# logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
#-----------------------------------------------------------------------------------------------#

import openpyxl
import re
import os
import subprocess
import smtplib
from shutil import copyfile


def process_credits():
	"""retrieve credit memos, discard any credits that reference 'FOR VERIFICATION ONLY'"""
	block = ""
	found = False
	verification_num = ""

	#open credit memo text file
	try:
		text_file = open("L:\\Merge and Print\\Print_Files\\" + agency_map[current[0]] + "\\" + current[0] + " Credits.txt")
		#parse through each line
		for line in text_file:
			if found:
				if WORD in line:
					output_file.write(block)
					block = ""
					global output
					output = True
					if search_string in line:
						block += line
					else:
						found = False
						break
				elif "invoice number:" in line.lower(): #captures invoice number
					invoice_number = int(line.lower().replace("invoice number:", ""))
					block += line
					if invoice_number == verification_num:
						block = ""
						found = False
				elif "for verification" in line.lower():
					verification_num = invoice_number
					block = ""
					found = False
					output = False
				else:
					block += line
			else:
				if search_string in line:
					found = True
					if output:
						block = line
					else:
						block = line.strip() + "\n"

		text_file.close()
	except:
		print("Print file for " + agency_map[current[0]] + " could not be found.")



def partner_credits(account_number, output_path, partner):
	"""Process Partner credits"""
	credit_file = os.path.join("L:", "Merge and Print", "Print_Files", partner, str(account_number)+".txt")
	if os.path.isfile(credit_file):
		copyfile(credit_file, output_path)
		global output
		output = True
	else:
		print(output_path)
		output = False



def update_workbook():
	"""Opens and updates any data connections contained within a specified workbook"""

	import win32com.client
	import os

	def close_excel_by_force(xl):
		"""forecfully closes excel"""
		import win32process
		import win32gui
		import win32api
		import win32con
		import time

		#Get the window's process id's
		hwnd = xl.Hwnd
		t,p = win32process.GetWindowThreadProcessId(hwnd)
		#Ask window nicely to close
		win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
		#Allow some time for the app to close
		time.sleep(3)
		#If the application didn't close, force close
		try:
			handle = win32api.OpenProcess(win32con.PROCESS_TERMINATE, 0, p)
			if handle:
				win32api.TerminateProcess(handle,0)
				win32api.CloseHandle(handle)
		except:
			pass


	filename = "L:\\Rite_Aid_Credit_Circulation\\Rite_Aid_Store_List.xlsx"
	xl = win32com.client.DispatchEx("Excel.Application")
	wb = xl.workbooks.open(filename)

	#makes workbook visible, currently disabeled because we're not
	#editing any information in the workbook
	xl.Visible = False
	xl.DisplayAlerts = False

	wb.RefreshAll()
	wb.Save()
	xl.Quit()
	xl.Application.Quit()
	close_excel_by_force(xl)

def convert_document(filepath):
	"""Coverts current txt document to PDF"""
	subprocess.run(["C:\\Program Files (x86)\\Text2PDF v1.5\\txt2pdf.exe",filepath,"-pfs8","-plm25","-prm25", \
				"-ptm25","-pbm25","-pdn:Lucida Console"])
	os.remove(filepath)

#
#
#       TO DO - Figure out how to format and send email.
#
#

####Main####

#Update the Rite Aid dealer listing
update_workbook()

WORD = "PRIMARY SEQ#:"

#agency map
agency_map = {"Atlanta" : "US_East",
		"Jackson" : "US_Midwest",
		"Specialty" : "Specialty",
		"Texas" : "US_Central",
		"Alaska" : "Alaska",
		"Fife" : "US_West",
		"Sacramento" : "US_West",
		"Salt Lake City" : "US_West",
		"Arizona" : "US_West",
		"Hawaii" : "US_West",
		"Benjamin" : "Benjamin",
		"Cowley" : "Cowley",
		"Kent" : "Kent"
		}

#open dealer document list
wb = openpyxl.load_workbook("L:\\Rite_Aid_Credit_Circulation\\Rite_Aid_Store_List.xlsx")
sheet = wb.active

dealer_count = 0
for i in range(2,sheet.max_row+1):
	dealer_count += 1
	output = False
	current = []
	for a in range(1,4):
		current = current + [sheet.cell(row = i, column = a).value]

	print(current,dealer_count)

	#account number, including blank characters, must be 12 digits long
	account_number_str = str(current[2])
	while len(account_number_str)<12:
		account_number_str = " " + account_number_str

	search_string = WORD + account_number_str

	#create output file for processing documents
	path = "L:\\Merge and Print\\In_Process\\" + agency_map[current[0]] + "\\" + current[0] + \
			" " + str(current[2]) + " " + re.sub('[^A-Za-z0-9]+',' ',str(current[1])) + ".txt"

	output_file = open(path,"a")

	#Begin processing documents
	if agency_map[current[0]] in ["Kent", "Cowley", "Benjamin"]:
		output_file.close() #Don't need the output file
		os.remove(path)
		partner_credits(current[2], path, current[0])
	else:
		process_credits()

	output_file.close()

	#Convert current document to PDF
	if output == False:
		try:
			os.remove(path)
		except:
			print("output file could not be found")

	else:
		convert_document(path)


#Temporary file created by Automation Anywhere - delete once Python has finished processing
#credits - this will allow AA to move on to the email portion of the script
try:
	os.remove("L:\\Rite_Aid_Credit_Circulation\\Script_Running.txt")
except:
	print("Temp file was not found")
