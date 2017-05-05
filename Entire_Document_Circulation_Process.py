#! python3
#Document Circulation
#Created by Curtis Cholon
#04/24/2017
"""Processes computac documents pulled from the Merge and Print Processes.
Documents are retrieved, formatted and sent via email or fax"""

import openpyxl
import re
import os
import subprocess

def stmts():
	"""Retrievs statements"""
	block = ""
	found = False
	
	text_file = open("L:\\Merge and Print\\Print_Files\\" + agency_map[current[0]] + "\\" + current[0] + " Statements.txt")
	#print(text_file)
	
	for line in text_file:
		if found:
			if WORD in line:
				if search_string in line:
					block += line
				else:
					found = False
					break
			else:
				block += line
		else:
			if search_string in line:
				found = True
				global output
				if output:
					block = line
				else:
					block = line.strip() + "\n"
					output = True
						
	output_file.write(block)
	#close the text file
	text_file.close()
	
def process_invoices():
	"""Retrieve Invoices"""
	block = ""
	found = False
	#open invoice text file
	text_file = open("L:\\Merge and Print\\Print_Files\\" + agency_map[current[0]] + "\\" + current[0] + " Invoices.txt")
	#print(text_file)

	#Parse through each line of te invoice file
	for line in text_file:
		if found:
			if WORD in line:
				if search_string in line:
					block += line
				else:
					found = False
					break
			else:
				block += line
		else:
			if search_string in line:
				found = True
				global output
				if output:
					block = line
				else:
					block = line.strip() + "\n"
					output = True
						
	#write contents to output file
	output_file.write(block)
	#close the invoice file
	text_file.close()
	
def process_credits():
	"""retrieve credit memos, discard any credits that reference 'FOR VERIFICATION ONLY'"""
	block = ""
	found = False
	verification_num = ""
	
	#open credit memo text file
	text_file = open("L:\\Merge and Print\\Print_Files\\" + agency_map[current[0]] + "\\" + current[0] + " Credits.txt")
	#print(text_file)
	
	#parse through each line
	for line in text_file:
		if found:
			if WORD in line:
				output_file.write(block)
				global output
				output = True
				if search_string in line:
					block += line
				else:
					found = False
					break
			elif "for verification" in line.lower():
				verification_num = invoice_number
				block = ""
				found = False
				output = False
			else:
				block += line
		else:
			if search_string in line:
				found = True
				if output:
					block = line
				else:
					block = line.strip() + "\n"
					
	text_file.close()
	
	
####Main####

WORD = "PRIMARY SEQ#:"

#agency map
agency_map = {"Atlanta" : "US_East",
		"Jackson" : "US_Midwest",
		"Specialty" : "Specialty",
		"Texas" : "US_Central",
		"Alaska" : "Alaska",
		"Fife" : "US_West",
		"Sacramento" : "US_West",
		"Salt Lake City" : "US_West",
		"Arizona" : "US_West",
		"Hawaii" : "US_West"
		}
		
#open dealer document list
wb = openpyxl.load_workbook("L:\\Merge and Print\\Customer_Documents - Original.xlsm")
sheet = wb.active

for i in range(2,sheet.max_row):
	output = False
	current = []
	for a in range(1,9):
		current = current + [sheet.cell(row = i, column = a).value]

		
	#account number, including blank characters, must be 12 digits long
	account_number_str = str(current[2])
	while len(account_number_str)<12:
		account_number_str = " " + account_number_str
			
	search_string = WORD + account_number_str
		
	#create output file for processing documents
	path = "L:\\Merge and Print\\In_Process\\" + agency_map[current[0]] + "\\" + current[0] + \
			" " + str(current[2]) + " " + re.sub('[^A-Za-z0-9]+',' ',str(current[1])) + ".txt"
	
	output_file = open(path,"a")
	
	#Begin processing documents
	if current[3].lower() == "y":
		stmts()
		
	if current[4].lower() == "y":
		process_invoices()
		
	if current[5].lower() == "y":
		process_credits()
		
	output_file.close()
	
	
		
	
			
		
				